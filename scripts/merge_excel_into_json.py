"""Incrementally merge Story Hook Excel raw data into existing JSON files.

Never deletes or regenerates existing records. Only appends / fills missing fields.
"""

from __future__ import annotations

import json
import re
import uuid
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "src" / "data"

ROLE_MAP = {
    "main role": "Main Role",
    "support role": "Support Role",
    "guest role": "Guest Role",
    "main": "Main Role",
    "support": "Support Role",
    "guest": "Guest Role",
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", clean(s)).casefold()


def parse_urls(text) -> list[str]:
    urls: list[str] = []
    for part in re.split(r"[\s,]+", clean(text)):
        part = part.strip()
        if part.startswith("http://") or part.startswith("https://"):
            urls.append(part)
    return urls


def parse_episode_links(text) -> list[dict]:
    text = clean(text)
    if not text:
        return []
    lines = [ln.strip() for ln in text.splitlines()]
    episodes: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line:
            i += 1
            continue
        if line.startswith("http://") or line.startswith("https://"):
            episodes.append({"title": "", "description": "", "link": line})
            i += 1
            continue
        title = line
        link = ""
        j = i + 1
        while j < len(lines) and not lines[j]:
            j += 1
        if j < len(lines) and (
            lines[j].startswith("http://") or lines[j].startswith("https://")
        ):
            link = lines[j]
            i = j + 1
        else:
            i += 1
        if link:
            episodes.append({"title": title, "description": "", "link": link})
    return episodes


def parse_networks(text) -> list[str]:
    names: list[str] = []
    for part in re.split(r"[,;/|]+", clean(text)):
        name = part.strip()
        if name:
            names.append(name)
    return names


def parse_cast_blocks(text) -> list[dict]:
    text = clean(text)
    if not text:
        return []
    blocks = re.split(r"\n\s*\n", text)
    casts: list[dict] = []
    for block in blocks:
        lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        if not lines:
            continue
        image = ""
        if lines[-1].startswith("http://") or lines[-1].startswith("https://"):
            image = lines.pop()
        if len(lines) < 2:
            continue
        name = lines[0]
        role = "Support Role"
        character_lines: list[str]
        if len(lines) >= 3 and norm_name(lines[-1]) in ROLE_MAP:
            role = ROLE_MAP[norm_name(lines[-1])]
            character_lines = lines[1:-1]
        else:
            character_lines = lines[1:]
        casts.append(
            {
                "name": name,
                "characterName": " ".join(character_lines).strip(),
                "role": role,
                "image": image,
            }
        )
    return casts


def load_json(path: Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, data, indent: int = 2) -> None:
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=indent)
        f.write("\n")


def find_story(stories, title: str, mm_title: str, watch_link: str):
    t = norm_name(title)
    m = norm_name(mm_title)
    w = clean(watch_link).rstrip("/").casefold()
    for story in stories:
        if t and norm_name(story.get("title", "")) == t:
            return story
        if m and norm_name(story.get("mmTitle", "")) == m:
            return story
        sw = clean(story.get("watchLink", "")).rstrip("/").casefold()
        if w and sw and w == sw:
            return story
    return None


def fill_missing(story: dict, field: str, value) -> bool:
    if value is None or value == "" or value == []:
        return False
    cur = story.get(field)
    if cur is None or cur == "" or cur == []:
        story[field] = value
        return True
    return False


def merge_photos(story: dict, photos: list[str]) -> int:
    existing = story.setdefault("photos", [])
    seen = set(existing)
    added = 0
    for photo in photos:
        if photo and photo not in seen:
            existing.append(photo)
            seen.add(photo)
            added += 1
    return added


def merge_episodes(story: dict, episodes: list[dict]) -> int:
    existing = story.setdefault("episodeLinks", [])
    seen = {
        (
            norm_name(ep.get("title", "")),
            clean(ep.get("link", "")).rstrip("/").casefold(),
        )
        for ep in existing
    }
    added = 0
    for ep in episodes:
        key = (
            norm_name(ep.get("title", "")),
            clean(ep.get("link", "")).rstrip("/").casefold(),
        )
        if key in seen:
            continue
        link_key = key[1]
        if link_key and any(
            clean(e.get("link", "")).rstrip("/").casefold() == link_key
            for e in existing
        ):
            continue
        existing.append(ep)
        seen.add(key)
        added += 1
    return added


def merge_cast_refs(story: dict, cast_refs: list[dict]) -> int:
    existing = story.setdefault("cast", [])
    seen = {
        (c.get("castUuid"), norm_name(c.get("characterName", ""))) for c in existing
    }
    added = 0
    for ref in cast_refs:
        key = (ref["castUuid"], norm_name(ref.get("characterName", "")))
        if key in seen:
            continue
        existing.append(ref)
        seen.add(key)
        added += 1
    return added


def merge_networks(story: dict, network_uuids: list[str]) -> int:
    existing = story.setdefault("orginalNetworks", [])
    seen = set(existing)
    added = 0
    for uid in network_uuids:
        if uid not in seen:
            existing.append(uid)
            seen.add(uid)
            added += 1
    return added


def merge(excel_path: Path) -> dict:
    stories = load_json(DATA / "stories.json")
    casts = load_json(DATA / "casts.json")
    networks = load_json(DATA / "original_network.json")

    orig_story_uuids = {s["uuid"] for s in stories}
    orig_cast_uuids = {c["uuid"] for c in casts}
    orig_network_uuids = {n["uuid"] for n in networks}
    orig_story_count = len(stories)
    orig_cast_count = len(casts)
    orig_network_count = len(networks)

    cast_by_name = {norm_name(c["name"]): c for c in casts}
    network_by_name = {norm_name(n["name"]): n for n in networks}

    def get_or_create_cast(name: str, image: str = "") -> tuple[str, bool]:
        key = norm_name(name)
        if key in cast_by_name:
            existing = cast_by_name[key]
            if image and not existing.get("image"):
                existing["image"] = image
            return existing["uuid"], False
        new = {"uuid": str(uuid.uuid4()), "name": name}
        if image:
            new["image"] = image
        casts.append(new)
        cast_by_name[key] = new
        return new["uuid"], True

    def get_or_create_network(name: str) -> tuple[str, bool]:
        key = norm_name(name)
        if key in network_by_name:
            return network_by_name[key]["uuid"], False
        new = {"uuid": str(uuid.uuid4()), "name": name}
        networks.append(new)
        network_by_name[key] = new
        return new["uuid"], True

    stats = {
        "rows": 0,
        "stories_created": 0,
        "stories_updated": 0,
        "casts_created": 0,
        "networks_created": 0,
        "photos_added": 0,
        "episodes_added": 0,
        "cast_refs_added": 0,
        "created_titles": [],
        "updated_titles": [],
        "new_casts": [],
        "new_networks": [],
    }

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, c).value for c, h in enumerate(headers, 1)}
        title = clean(row.get("Title"))
        if not title:
            continue
        stats["rows"] += 1

        mm_title = clean(row.get("MMTitle"))
        story_text = clean(row.get("Story"))
        country = clean(row.get("Country"))
        rating = clean(row.get("Rating"))
        stype = clean(row.get("Type")) or "Drama"
        sformat = clean(row.get("Format")) or "Standard Series"
        watch_link = clean(row.get("WatchLink"))
        episodes_raw = row.get("Epiosodes")
        try:
            episodes_n = (
                int(float(episodes_raw)) if episodes_raw not in (None, "") else 0
            )
        except (TypeError, ValueError):
            episodes_n = 0
        episode_links = parse_episode_links(row.get("EpiosodeLinks"))
        aired = clean(row.get("Aired"))
        duration = clean(row.get("Duration"))
        network_names = parse_networks(row.get("Orginal Networks"))
        cast_blocks = parse_cast_blocks(
            row.get("CastName,CharacterName and CastImage")
        )
        photos = parse_urls(row.get("Photos"))
        cover = clean(row.get("Cover Photo"))

        network_uuids: list[str] = []
        for nname in network_names:
            uid, created = get_or_create_network(nname)
            network_uuids.append(uid)
            if created:
                stats["networks_created"] += 1
                stats["new_networks"].append(nname)

        cast_refs: list[dict] = []
        for cb in cast_blocks:
            uid, created = get_or_create_cast(cb["name"], cb.get("image", ""))
            if created:
                stats["casts_created"] += 1
                stats["new_casts"].append(cb["name"])
            cast_refs.append(
                {
                    "castUuid": uid,
                    "characterName": cb["characterName"],
                    "role": cb["role"],
                }
            )

        existing = find_story(stories, title, mm_title, watch_link)
        if existing is None:
            stories.append(
                {
                    "uuid": str(uuid.uuid4()),
                    "title": title,
                    "mmTitle": mm_title,
                    "story": story_text,
                    "country": country,
                    "rating": rating,
                    "type": stype,
                    "format": sformat,
                    "watchLink": watch_link,
                    "episodes": episodes_n,
                    "episodeLinks": episode_links,
                    "aired": aired,
                    "duration": duration,
                    "orginalNetworks": network_uuids,
                    "cast": cast_refs,
                    "photos": photos,
                    "coverPhoto": cover,
                }
            )
            stats["stories_created"] += 1
            stats["created_titles"].append(title)
        else:
            stats["stories_updated"] += 1
            stats["updated_titles"].append(existing.get("title", title))
            fill_missing(existing, "mmTitle", mm_title)
            fill_missing(existing, "story", story_text)
            fill_missing(existing, "country", country)
            fill_missing(existing, "rating", rating)
            fill_missing(existing, "type", stype)
            fill_missing(existing, "format", sformat)
            fill_missing(existing, "watchLink", watch_link)
            if episodes_n and not existing.get("episodes"):
                existing["episodes"] = episodes_n
            fill_missing(existing, "aired", aired)
            fill_missing(existing, "duration", duration)
            if cover and not existing.get("coverPhoto"):
                existing["coverPhoto"] = cover
            stats["photos_added"] += merge_photos(existing, photos)
            stats["episodes_added"] += merge_episodes(existing, episode_links)
            stats["cast_refs_added"] += merge_cast_refs(existing, cast_refs)
            merge_networks(existing, network_uuids)

    assert orig_story_uuids.issubset({s["uuid"] for s in stories})
    assert orig_cast_uuids.issubset({c["uuid"] for c in casts})
    assert orig_network_uuids.issubset({n["uuid"] for n in networks})

    write_json(DATA / "stories.json", stories, indent=2)
    write_json(DATA / "casts.json", casts, indent=4)
    write_json(DATA / "original_network.json", networks, indent=4)

    stats.update(
        {
            "stories_before": orig_story_count,
            "casts_before": orig_cast_count,
            "networks_before": orig_network_count,
            "stories_total": len(stories),
            "casts_total": len(casts),
            "networks_total": len(networks),
        }
    )
    return stats


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "excel",
        nargs="?",
        default=str(
            Path.home() / "Downloads" / "Story Hook Raw Data.xlsx"
        ),
        help="Path to Story Hook Raw Data.xlsx",
    )
    args = parser.parse_args()
    result = merge(Path(args.excel))
    print(json.dumps(result, ensure_ascii=False, indent=2))
